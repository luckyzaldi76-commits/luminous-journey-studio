from pathlib import Path
from openpyxl import load_workbook

DATABASE = Path(r"E:\LUMINOUS JOURNEY\01_DATABASE\TGL.xlsx")


def find_reading(search_date):

    workbook = load_workbook(DATABASE)

    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        tgl = row[0]
        bacaan1 = row[1]
        bacaan2 = row[2]
        injil = row[3]

        if str(tgl) == search_date:

            print("=" * 50)
            print("LUMINOUS JOURNEY STUDIO")
            print("=" * 50)

            print("Tanggal   :", tgl)
            print("Bacaan I  :", bacaan1)
            print("Bacaan II :", bacaan2)
            print("Injil     :", injil)

            workbook.close()
            return

    print("Tanggal tidak ditemukan.")

    workbook.close()


if __name__ == "__main__":

    find_reading("2026-08-03")