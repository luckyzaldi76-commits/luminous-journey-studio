from dataclasses import dataclass
from datetime import datetime

from openpyxl import load_workbook

from config import DATABASE_DIR


DATABASE = DATABASE_DIR / "TGL.xlsx"


@dataclass
class Reading:

    date: str
    reading1: str
    reading2: str
    gospel: str


class ReadingFinder:

    def __init__(self):

        self.database = DATABASE

    def find(self, search_date):

        workbook = load_workbook(self.database)

        sheet = workbook.active

        try:

            for row in sheet.iter_rows(min_row=2, values_only=True):

                excel_date = row[0]

                if isinstance(excel_date, datetime):

                    date = excel_date.strftime("%Y-%m-%d")

                else:

                    date = str(excel_date).strip()

                print(f"Excel : {date} | Input : {search_date}")

                if date != search_date:
                    continue

                return Reading(

                    date=date,

                    reading1=row[1] if row[1] else "",

                    reading2=row[2] if row[2] else "",

                    gospel=row[3] if row[3] else ""

                )

        finally:

            workbook.close()

        return None


finder = ReadingFinder()


def find_reading(search_date):

    return finder.find(search_date)


if __name__ == "__main__":

    reading = find_reading("2026-07-31")

    if reading:

        print()
        print(reading)

    else:

        print("Reading Not Found")