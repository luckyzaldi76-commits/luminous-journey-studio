from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class GospelInput:

    reference: str
    text: str
    language: str = "IND"


class GospelInputService:

    def __init__(self, database_file=None):

        self.database_file = Path(
            database_file
            or "01_DATABASE/TGL.xlsx"
        )

    def validate(self, gospel: str) -> str:

        if not isinstance(gospel, str):

            raise TypeError(
                "Gospel must be a string."
            )

        gospel = gospel.strip()

        if not gospel:

            raise ValueError(
                "Gospel cannot be empty."
            )

        return gospel

    def create(
        self,
        reference: str,
        text: str,
        language: str = "IND",
    ) -> GospelInput:

        reference = self.validate(reference)
        text = self.validate(text)
        language = self.validate(language)

        return GospelInput(
            reference=reference,
            text=text,
            language=language.upper(),
        )

    def find(
        self,
        reference: str,
    ) -> Optional[GospelInput]:

        reference = self.validate(reference)

        if not self.database_file.exists():

            raise FileNotFoundError(
                self.database_file
            )

        workbook = load_workbook(
            self.database_file,
            read_only=True,
            data_only=True,
        )

        try:

            for sheet in workbook.worksheets:

                rows = sheet.iter_rows(
                    values_only=True
                )

                headers = next(
                    rows,
                    None,
                )

                if not headers:
                    continue

                normalized = [
                    str(value).strip().lower()
                    if value is not None
                    else ""
                    for value in headers
                ]

                reference_index = None
                text_index = None
                language_index = None

                for index, header in enumerate(
                    normalized
                ):

                    if header in {
                        "reference",
                        "ref",
                        "gospel",
                        "bacaan",
                    }:
                        reference_index = index

                    if header in {
                        "text",
                        "teks",
                        "gospel_text",
                        "isi",
                    }:
                        text_index = index

                    if header in {
                        "language",
                        "lang",
                        "bahasa",
                    }:
                        language_index = index

                if (
                    reference_index is None
                    or text_index is None
                ):
                    continue

                for row in rows:

                    if len(row) <= reference_index:
                        continue

                    value = row[
                        reference_index
                    ]

                    if value is None:
                        continue

                    if (
                        str(value).strip().lower()
                        == reference.lower()
                    ):

                        text = row[text_index]

                        if text is None:
                            return None

                        language = "IND"

                        if (
                            language_index is not None
                            and len(row) > language_index
                            and row[language_index]
                        ):
                            language = str(
                                row[language_index]
                            ).strip().upper()

                        return GospelInput(
                            reference=str(value).strip(),
                            text=str(text).strip(),
                            language=language,
                        )

        finally:

            workbook.close()

        return None
