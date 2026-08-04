from pathlib import Path
from openpyxl import load_workbook
from models.reading import Reading

DATABASE = Path(r"E:\LUMINOUS JOURNEY\01_DATABASE")
FILE = DATABASE / "TGL.xlsx"


def get_reading(search_date):

    if not FILE.exists():
        return None

    workbook = load_workbook(FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):

        tgl = row[0]
        bacaan1 = row[1]
        bacaan2 = row[2]
        injil = row[3]

        # Ubah tanggal Excel menjadi YYYY-MM-DD
        if hasattr(tgl, "strftime"):
            tgl = tgl.strftime("%Y-%m-%d")
        else:
            tgl = str(tgl)

        if tgl == search_date:

            workbook.close()

            return Reading(
                date=tgl,
                reading1=bacaan1,
                reading2=bacaan2,
                gospel=injil
            )

    workbook.close()
    return None